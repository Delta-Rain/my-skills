#!/usr/bin/env python3
"""Compare two .xlsx sheets and report changes in the 海外简中 column
(column E, 1-based index 5). Designed for the tdoc-sheet-sync workflow.

It reports:
  - 增量 (increments): rows present in NEW but not OLD. Because alignment is
    done by a content key (all columns EXCEPT E), inserts that appear in the
    MIDDLE of the document are still detected correctly.
  - 改量 (modifications): rows that exist in both but whose E value changed.
  - 减量 (removals): rows present in OLD but not NEW (reported for awareness).

Usage:
  python diff_overseas_column.py --old 平面文案.xlsx --new 平面文案.new.xlsx
  python diff_overseas_column.py --old a.xlsx --new b.xlsx --sheet 平面文案 --report report.md
"""
import argparse
import difflib
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

E_COL = 5  # 1-based: A=1,B=2,C=3,D=4,E=5  -> 海外简中
NAME_COL = 2  # B: 平面名
CN_COL = 4    # D: 国服简中


def load_rows(path, sheet):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    wb.close()
    return ws.title, rows


def key_of(row):
    # stable key = all columns except E, joined; ignores E so pure E edits
    # stay "equal" and are caught separately as 改量.
    parts = []
    for i, v in enumerate(row, start=1):
        if i == E_COL:
            continue
        parts.append("" if v is None else str(v))
    return "\u0001".join(parts)


def cell(row, col):
    if col - 1 < len(row):
        return row[col - 1]
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--old", required=True)
    ap.add_argument("--new", required=True)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--report", default=None, help="optional path to write a markdown report")
    ap.add_argument("--e-col", type=int, default=E_COL)
    args = ap.parse_args()

    e = args.e_col
    old_sheet, old_rows = load_rows(args.old, args.sheet)
    new_sheet, new_rows = load_rows(args.new, args.sheet)

    old_keys = [key_of(r) for r in old_rows]
    new_keys = [key_of(r) for r in new_rows]

    increments, modifies, removals = [], [], []

    sm = difflib.SequenceMatcher(None, old_keys, new_keys, autojunk=False)
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            for k in range(i2 - i1):
                o_row = old_rows[i1 + k]
                n_row = new_rows[j1 + k]
                o_e = cell(o_row, e)
                n_e = cell(n_row, e)
                if o_e != n_e:
                    modifies.append({
                        "name": cell(n_row, NAME_COL),
                        "cn": cell(n_row, CN_COL),
                        "old": o_e, "new": n_e,
                    })
        elif tag == "insert":
            for j in range(j1, j2):
                n_row = new_rows[j]
                increments.append({
                    "name": cell(n_row, NAME_COL),
                    "cn": cell(n_row, CN_COL),
                    "e": cell(n_row, e),
                })
        elif tag == "delete":
            for i in range(i1, i2):
                o_row = old_rows[i]
                removals.append({
                    "name": cell(o_row, NAME_COL),
                    "cn": cell(o_row, CN_COL),
                    "e": cell(o_row, e),
                })
        elif tag == "replace":
            # new side = increments, old side = removals; also flag E changes
            for j in range(j1, j2):
                n_row = new_rows[j]
                increments.append({
                    "name": cell(n_row, NAME_COL),
                    "cn": cell(n_row, CN_COL),
                    "e": cell(n_row, e),
                })
            for i in range(i1, i2):
                o_row = old_rows[i]
                removals.append({
                    "name": cell(o_row, NAME_COL),
                    "cn": cell(o_row, CN_COL),
                    "e": cell(o_row, e),
                })

    lines = []
    lines.append("# 海外简中（E列）增量 / 改量 对比报告")
    lines.append("")
    lines.append("- 对比文件：旧 `%s`（%s） vs 新 `%s`（%s）" % (args.old, old_sheet, args.new, new_sheet))
    lines.append("- 总行数：旧 %d → 新 %d" % (len(old_rows), len(new_rows)))
    lines.append("- 改量（数值变化）：%d 处" % len(modifies))
    lines.append("- 增量（新增行）：%d 处" % len(increments))
    lines.append("- 减量（减少行）：%d 处" % len(removals))
    lines.append("")

    if modifies:
        lines.append("## 改量（海外简中内容被修改）")
        for m in modifies:
            lines.append("- **%s**（国服简中：%s）" % (_s(m["name"]), _s(m["cn"])))
            lines.append("  - 旧：%s" % _s(m["old"]))
            lines.append("  - 新：%s" % _s(m["new"]))
        lines.append("")

    if increments:
        lines.append("## 增量（新增行，可能插入在原文中间）")
        for inc in increments:
            lines.append("- **%s**（国服简中：%s）海外简中：`%s`" % (_s(inc["name"]), _s(inc["cn"]), _s(inc["e"])))
        lines.append("")

    if removals:
        lines.append("## 减量（被删除的行，供参考）")
        for rm in removals:
            lines.append("- **%s**（国服简中：%s）海外简中：`%s`" % (_s(rm["name"]), _s(rm["cn"]), _s(rm["e"])))
        lines.append("")

    text = "\n".join(lines)
    print(text)
    if args.report:
        with open(args.report, "w", encoding="utf-8") as f:
            f.write(text)
        print("[done] report written:", args.report)


def _s(v):
    if v is None:
        return "（空）"
    s = str(v).replace("\n", " ⏎ ")
    return s if s.strip() else "（空）"


if __name__ == "__main__":
    main()
